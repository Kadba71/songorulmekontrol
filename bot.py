import asyncio
import logging
import os
import tempfile
from datetime import date, datetime, time, timedelta, timezone
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

from dotenv import load_dotenv
from openpyxl import load_workbook
from telegram import Update
from telegram.ext import Application, CommandHandler, ContextTypes, MessageHandler, filters
from telethon import TelegramClient
from telethon.sessions import StringSession
from telethon.tl.types import (
    UserStatusEmpty,
    UserStatusLastMonth,
    UserStatusLastWeek,
    UserStatusOffline,
    UserStatusOnline,
    UserStatusRecently,
)

import database

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s: %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("lastseen-bot")

load_dotenv()

BOT_TOKEN = os.getenv("BOT_TOKEN", "")
ALLOWED_CHAT_IDS = {
    int(x.strip())
    for x in os.getenv("ALLOWED_CHAT_IDS", "").split(",")
    if x.strip()
}
ALLOWED_GROUP_TITLES = {
    x.strip().lower()
    for x in os.getenv("ALLOWED_GROUP_TITLES", "").split(",")
    if x.strip()
}
ALERT_CHAT_ID = int(os.getenv("ALERT_CHAT_ID", "0") or 0)
DEFAULT_THRESHOLD_MINUTES = int(os.getenv("DEFAULT_THRESHOLD_MINUTES", "15"))
ALERT_COOLDOWN_MINUTES = int(os.getenv("ALERT_COOLDOWN_MINUTES", "30"))
CHECK_INTERVAL_SECONDS = int(os.getenv("CHECK_INTERVAL_SECONDS", "60"))
MONITOR_START_HOUR = int(os.getenv("MONITOR_START_HOUR", "9"))
MONITOR_START_MINUTE = int(os.getenv("MONITOR_START_MINUTE", "45"))
MONITOR_END_HOUR = int(os.getenv("MONITOR_END_HOUR", "19"))
MONITOR_END_MINUTE = int(os.getenv("MONITOR_END_MINUTE", "0"))
DAILY_REPORT_HOUR = int(os.getenv("DAILY_REPORT_HOUR", "19"))
DAILY_REPORT_MINUTE = int(os.getenv("DAILY_REPORT_MINUTE", "30"))
DAILY_REPORT_CHECK_INTERVAL_SECONDS = int(os.getenv("DAILY_REPORT_CHECK_INTERVAL_SECONDS", "60"))
APP_TIMEZONE_NAME = os.getenv("APP_TIMEZONE", "Europe/Istanbul").strip() or "Europe/Istanbul"

TELEGRAM_API_ID = int(os.getenv("TELEGRAM_API_ID", "0") or 0)
TELEGRAM_API_HASH = os.getenv("TELEGRAM_API_HASH", "")
TELETHON_STRING_SESSION = os.getenv("TELETHON_STRING_SESSION", "")

telethon_client: TelegramClient | None = None
last_authorized_chat_id: int = 0
DAILY_SUMMARY_LAST_SENT_KEY = "daily_summary_last_sent_date"

def resolve_app_timezone(name: str):
    normalized = (name or "").strip()
    if not normalized:
        return timezone(timedelta(hours=3))

    try:
        return ZoneInfo(normalized)
    except ZoneInfoNotFoundError:
        if normalized in {"Europe/Istanbul", "Turkey Standard Time"}:
            logger.warning(
                "APP_TIMEZONE '%s' sistemde bulunamadı. UTC+03:00 kullanılacak.",
                normalized,
            )
            return timezone(timedelta(hours=3))

        logger.warning(
            "Geçersiz APP_TIMEZONE: %s. UTC kullanılacak.",
            normalized,
        )
        return timezone.utc


APP_TIMEZONE = resolve_app_timezone(APP_TIMEZONE_NAME)


def is_authorized_chat(chat_id: int, chat_title: str | None) -> bool:
    if ALLOWED_CHAT_IDS and chat_id not in ALLOWED_CHAT_IDS:
        return False
    if ALLOWED_GROUP_TITLES:
        normalized_title = (chat_title or "").strip().lower()
        if normalized_title not in ALLOWED_GROUP_TITLES:
            return False
    return True


def parse_csv_args(text: str) -> list[str]:
    if not text:
        return []
    return [x.strip() for x in text.split(",") if x.strip()]


def parse_hour_text(text: str) -> int | None:
    text = text.strip().lower()
    if not text:
        return None
    tokens = text.split()
    if not tokens:
        return None
    try:
        return int(tokens[0])
    except ValueError:
        return None


def parse_hhmm(text: str) -> time | None:
    raw = text.strip()
    parts = raw.split(":")
    if len(parts) != 2:
        return None
    try:
        hour = int(parts[0])
        minute = int(parts[1])
    except ValueError:
        return None
    if not (0 <= hour <= 23 and 0 <= minute <= 59):
        return None
    return time(hour=hour, minute=minute)


def should_skip_for_break_window(now_dt: datetime, break_start_hhmm: str, break_end_hhmm: str) -> bool:
    start_time = parse_hhmm(break_start_hhmm)
    end_time = parse_hhmm(break_end_hhmm)
    if start_time is None or end_time is None:
        return False
    if start_time >= end_time:
        return False

    break_start = datetime.combine(now_dt.date(), start_time, tzinfo=now_dt.tzinfo)
    break_resume = datetime.combine(now_dt.date(), end_time, tzinfo=now_dt.tzinfo) + timedelta(minutes=10)
    return break_start <= now_dt < break_resume


async def db_call(func, *args, **kwargs):
    return await asyncio.to_thread(func, *args, **kwargs)


def normalize_weekday(text: str) -> str:
    value = text.strip().lower()
    aliases = {
        "pazartesi": "pazartesi",
        "salı": "salı",
        "sali": "salı",
        "çarşamba": "çarşamba",
        "carsamba": "çarşamba",
        "perşembe": "perşembe",
        "persembe": "perşembe",
        "cuma": "cuma",
        "cumartesi": "cumartesi",
        "pazar": "pazar",
        "monday": "pazartesi",
        "tuesday": "salı",
        "wednesday": "çarşamba",
        "thursday": "perşembe",
        "friday": "cuma",
        "saturday": "cumartesi",
        "sunday": "pazar",
    }
    return aliases.get(value, value)


def today_weekday_tr() -> str:
    names = [
        "pazartesi",
        "salı",
        "çarşamba",
        "perşembe",
        "cuma",
        "cumartesi",
        "pazar",
    ]
    return names[get_now_local().weekday()]


def should_skip_for_department_weekly_off(
    department_name: str | None,
    weekly_off_day: str | None,
    weekday_name: str | None = None,
) -> bool:
    if not (department_name or "").strip():
        return False
    normalized_weekly_off = normalize_weekday(weekly_off_day or "")
    if not normalized_weekly_off:
        return False
    return normalized_weekly_off == (weekday_name or today_weekday_tr())


def get_now_local() -> datetime:
    return datetime.now(APP_TIMEZONE)


def get_today_local_iso() -> str:
    return get_now_local().date().isoformat()


def format_responsible(username: str | None) -> str:
    if not username or username == "-":
        return "-"
    return f"@{username}"


def build_help_text() -> str:
    return (
        "Komutlar ve açıklamalar:\n"
        "/start -> Yardım metnini gösterir.\n"
        "/yardim -> Yardım metnini gösterir.\n"
        "/help -> Yardım metnini gösterir.\n\n"
        "/sure (dakika), (departman) -> Departman eşik süresini ayarlar.\n"
        "Örn: /sure 20, satısekibi1\n\n"
        "/sureguncelle (dakika), (departman) -> Departman süresini günceller.\n\n"
        "/personelekle (personel @), (sorumlu @), (departman) -> Personeli takibe ekler.\n"
        "Örn: /personelekle @ahmet_taha, @ayse_su, satısekibi2\n\n"
        "/silpersonel (personel @) -> Personeli siler.\n\n"
        "/eklesorumlu (sorumlu @), (departman) -> Departmana sorumlu ekler.\n"
        "/silsorumlu (sorumlu @), (departman) -> Departmandan sorumlu çıkarır.\n\n"
        "/ekledepartman (departman) -> Departman ekler.\n"
        "/sildepartman (departman) -> Departman siler.\n\n"
        "/haftalikizin (departman), (gün) -> O gün departman kontrol edilmez.\n"
        "Örn: /haftalikizin satısekibi1, çarşamba\n\n"
        "/kontrolhaftalikizin -> Haftalık izin tanımlı departmanları listeler.\n\n"
        "/izin (personel @) -> Personel bugün kontrol edilmez.\n\n"
        "/kontrolizin -> Aktif personel izinlerini listeler.\n\n"
        "/iziniptal (personel @), (departman) -> Personelin tam gün iznini iptal eder.\n"
        "Örn: /iziniptal @ahmet_taha, satısekibi1\n\n"

        "/saatlikizin (personel @), (saat) -> Personel belirtilen saat boyunca kontrol edilmez.\n"
        "Örn: /saatlikizin @ahmet_taha, 2 saat\n\n"
        "/saatlikiziniptal (personel @), (departman) -> Personelin saatlik iznini iptal eder.\n"
        "Örn: /saatlikiziniptal @ahmet_taha, satısekibi1\n\n"
        "/mola (HH:MM), (HH:MM) -> Bu aralıkta kontrol durur, bitişten 10 dk sonra başlar.\n"
        "Örn: /mola 14:00, 15:00 (kontrol 15:10'da başlar)\n\n"
        "/yukle -> Excel dosyasını toplu personel ekleme için işler.\n"
        "Kullanım: Excel dosyasını gruba /yukle açıklamasıyla gönderin.\n\n"
        "/rapor (departman) -> İlgili departmanın bugünkü ihlal adetlerini listeler.\n"
        "Örn: /rapor satısekibi1\n\n"
        "/listele -> Kayıtlı personelleri departman bazında listeler.\n"
        "/chatid -> Bulunduğun sohbetin chat id bilgisini verir.\n"
        "Otomatik: 09:45-19:00 arası kontrol, 19:30 gün sonu ihlal raporu.\n"
        "Not: Komutlarda Türkçe karakter kullanmayın (ör: /kontrolhaftalikizin)."
    )


def _to_cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _is_header_row(personnel: str, responsible: str, department: str) -> bool:
    header_text = f"{personnel} {responsible} {department}".lower()
    return all(token in header_text for token in ("personel", "sorumlu", "departman"))


def _is_supported_excel_file(file_name: str) -> bool:
    return file_name.lower().endswith((".xlsx", ".xlsm"))


def is_within_monitor_hours(now_dt: datetime | None = None) -> bool:
    now_dt = now_dt or get_now_local()
    now_time = now_dt.time()
    start = time(MONITOR_START_HOUR, MONITOR_START_MINUTE)
    end = time(MONITOR_END_HOUR, MONITOR_END_MINUTE)
    return start <= now_time <= end


def resolve_target_chat_id() -> int:
    return ALERT_CHAT_ID if ALERT_CHAT_ID else 0


def _validate_range(name: str, value: int, minimum: int, maximum: int) -> None:
    if not (minimum <= value <= maximum):
        raise RuntimeError(f"{name} değeri {minimum}-{maximum} aralığında olmalı")


def validate_config() -> None:
    if not BOT_TOKEN:
        raise RuntimeError("BOT_TOKEN zorunlu")
    if not ALERT_CHAT_ID:
        raise RuntimeError("ALERT_CHAT_ID zorunlu")
    _validate_range("DEFAULT_THRESHOLD_MINUTES", DEFAULT_THRESHOLD_MINUTES, 1, 24 * 60)
    _validate_range("ALERT_COOLDOWN_MINUTES", ALERT_COOLDOWN_MINUTES, 1, 24 * 60)
    _validate_range("CHECK_INTERVAL_SECONDS", CHECK_INTERVAL_SECONDS, 10, 24 * 60 * 60)
    _validate_range("MONITOR_START_HOUR", MONITOR_START_HOUR, 0, 23)
    _validate_range("MONITOR_END_HOUR", MONITOR_END_HOUR, 0, 23)
    _validate_range("MONITOR_START_MINUTE", MONITOR_START_MINUTE, 0, 59)
    _validate_range("MONITOR_END_MINUTE", MONITOR_END_MINUTE, 0, 59)
    _validate_range("DAILY_REPORT_HOUR", DAILY_REPORT_HOUR, 0, 23)
    _validate_range("DAILY_REPORT_MINUTE", DAILY_REPORT_MINUTE, 0, 59)
    _validate_range("DAILY_REPORT_CHECK_INTERVAL_SECONDS", DAILY_REPORT_CHECK_INTERVAL_SECONDS, 30, 3600)


async def require_auth(update: Update) -> bool:
    global last_authorized_chat_id
    chat_id = update.effective_chat.id if update.effective_chat else 0
    chat_title = update.effective_chat.title if update.effective_chat else ""
    if not is_authorized_chat(chat_id, chat_title):
        await update.message.reply_text("Bu sohbet yetkili değil.")
        return False
    if chat_id:
        last_authorized_chat_id = chat_id
    return True


async def chatid_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    chat_id = update.effective_chat.id if update.effective_chat else 0
    await update.message.reply_text(f"Bu sohbetin chat id değeri: {chat_id}")


async def start_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    await update.message.reply_text(build_help_text())


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    await update.message.reply_text(build_help_text())


async def sure_set_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /sure 20, satısekibi1")
        return

    try:
        minutes = int(parts[0])
    except ValueError:
        await update.message.reply_text("Dakika değeri sayısal olmalı.")
        return

    if minutes < 1:
        await update.message.reply_text("Dakika en az 1 olmalı.")
        return

    department = parts[1]
    await db_call(database.set_department_threshold, department, minutes)
    await update.message.reply_text(f"Departman süresi ayarlandı: {department} -> {minutes} dakika")


async def sure_guncelle_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    await sure_set_cmd(update, context)


async def personelekle_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) < 3:
        await update.message.reply_text(
            "Kullanım: /personelekle @personel, @sorumlu, departman"
        )
        return

    personel, sorumlu, departman = parts[0], parts[1], parts[2]
    username = await db_call(database.add_personnel, personel, sorumlu, departman)
    await update.message.reply_text(
        f"Personel eklendi/güncellendi: @{username} | sorumlu: {sorumlu} | departman: {departman}"
    )


async def silpersonel_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2].strip()
    if not args_text:
        await update.message.reply_text("Kullanım: /silpersonel @personel")
        return
    ok = await db_call(database.remove_personnel, args_text)
    await update.message.reply_text("Personel silindi." if ok else "Personel bulunamadı.")


async def eklesorumlu_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /eklesorumlu @ayse_su, satısekibi1")
        return

    responsible = parts[0]
    department = parts[1]
    dep_name, username = await db_call(database.add_department_responsible, department, responsible)
    await update.message.reply_text(f"Departmana sorumlu eklendi: {dep_name} -> @{username}")


async def silsorumlu_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /silsorumlu @ayse_su, satısekibi1")
        return
    responsible = parts[0]
    department = parts[1]
    ok = await db_call(database.remove_department_responsible, department, responsible)
    await update.message.reply_text(
        "Sorumlu departmandan silindi." if ok else "Eşleşen sorumlu/departman bulunamadı."
    )


async def ekledepartman_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    name = update.message.text.partition(" ")[2].strip()
    if not name:
        await update.message.reply_text("Kullanım: /ekledepartman satısekibi1")
        return
    await db_call(database.add_department, name)
    await update.message.reply_text(f"Departman eklendi: {name}")


async def haftalikizin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /haftalikizin satısekibi1, çarşamba")
        return

    department = parts[0]
    weekday = normalize_weekday(parts[1])
    valid_days = {"pazartesi", "salı", "çarşamba", "perşembe", "cuma", "cumartesi", "pazar"}
    if weekday not in valid_days:
        await update.message.reply_text("Geçersiz gün adı. Örnek: pazartesi, çarşamba, cuma")
        return

    await db_call(database.set_department_weekly_off, department, weekday)
    await update.message.reply_text(f"Haftalık izin tanımlandı: {department} -> {weekday}")


async def kontrolhaftalikizin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return

    rows = await db_call(database.list_departments_with_weekly_off)
    if not rows:
        await update.message.reply_text("Haftalık izin tanımlı departman yok.")
        return

    lines = ["Haftalık izin tanımlı departmanlar:"]
    for row in rows:
        department_name = row["name"] or "-"
        weekly_off = normalize_weekday(row["weekly_off_day"] or "") or "-"
        lines.append(f"- {department_name}: {weekly_off}")

    await update.message.reply_text("\n".join(lines))


async def izin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    personnel = update.message.text.partition(" ")[2].strip()
    if not personnel:
        await update.message.reply_text("Kullanım: /izin @ahmet_taha")
        return

    today = get_today_local_iso()
    ok = await db_call(database.set_personnel_day_off_today, personnel, today)
    await update.message.reply_text(
        "Personel bugün izinli olarak işaretlendi." if ok else "Personel bulunamadı."
    )


async def saatlikizin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /saatlikizin @ahmet_taha, 2 saat")
        return

    personnel = parts[0]
    hours = parse_hour_text(parts[1])
    if hours is None or hours < 1:
        await update.message.reply_text("Saat değeri geçersiz. Örnek: 2 saat")
        return

    until_dt = datetime.now(timezone.utc) + timedelta(hours=hours)
    ok = await db_call(database.set_personnel_hourly_off, personnel, until_dt.isoformat())
    await update.message.reply_text(
        f"Personel {hours} saat izinli olarak işaretlendi." if ok else "Personel bulunamadı."
    )


async def saatlikiziniptal_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /saatlikiziniptal @ahmet_taha, satısekibi1")
        return

    personnel = parts[0]
    department = parts[1]
    ok = await db_call(database.cancel_personnel_hourly_off, personnel, department)
    await update.message.reply_text(
        "Personelin izni iptal edilmiştir. Son görülmesi kontrol edilecektir artık."
        if ok
        else "Aktif saatlik izin bulunamadı veya personel/departman eşleşmedi."
    )


async def iziniptal_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    args_text = update.message.text.partition(" ")[2]
    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /iziniptal @ahmet_taha, satısekibi1")
        return

    personnel = parts[0]
    department = parts[1]
    ok = await db_call(database.cancel_personnel_day_off, personnel, department)
    await update.message.reply_text(
        "Personelin tam gün izni iptal edilmiştir. Artık personelin son görülmesi takip edilecektir."
        if ok
        else "Aktif tam gün izin bulunamadı veya personel/departman eşleşmedi."
    )


async def kontrolizin_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return

    records = await db_call(database.list_personnel)
    if not records:
        await update.message.reply_text("Kayıtlı personel yok.")
        return

    today_iso = get_today_local_iso()
    now_utc = datetime.now(timezone.utc)
    full_day_lines: list[str] = []
    hourly_lines: list[str] = []

    for row in records:
        username = row["username"]
        department = row["department_name"] or "-"

        if row["day_off_date"] == today_iso:
            full_day_lines.append(f"@{username} ({department}) - bugün")
            continue

        exempt_until = row["exempt_until"]
        if not exempt_until:
            continue

        try:
            until_dt = datetime.fromisoformat(exempt_until)
        except ValueError:
            continue
        if until_dt.tzinfo is None:
            until_dt = until_dt.replace(tzinfo=timezone.utc)
        if until_dt <= now_utc:
            continue

        remaining_minutes = max(1, int((until_dt - now_utc).total_seconds() // 60))
        until_local = until_dt.astimezone(APP_TIMEZONE).strftime("%H:%M")
        hourly_lines.append(
            f"@{username} ({department}) - {remaining_minutes} dk kaldı (bitiş {until_local})"
        )

    if not full_day_lines and not hourly_lines:
        await update.message.reply_text("Aktif izinli personel yok.")
        return

    message_parts = ["Aktif izinli personeller:"]
    if full_day_lines:
        message_parts.append("\nTam gün izin:")
        message_parts.extend(f"- {line}" for line in sorted(full_day_lines))
    if hourly_lines:
        message_parts.append("\nSaatlik izin:")
        message_parts.extend(f"- {line}" for line in sorted(hourly_lines))

    await update.message.reply_text("\n".join(message_parts))


async def mola_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return

    args_text = update.message.text.partition(" ")[2].strip()
    if not args_text:
        current_start, current_end = await db_call(database.get_break_window)
        if current_start and current_end:
            await update.message.reply_text(
                f"Mevcut mola aralığı: {current_start}-{current_end} (kontrol {current_end} sonrası 10 dk bekler)."
            )
        else:
            await update.message.reply_text("Kullanım: /mola 14:00,15:00")
        return

    parts = parse_csv_args(args_text)
    if len(parts) != 2:
        await update.message.reply_text("Kullanım: /mola 14:00,15:00")
        return

    start_text, end_text = parts[0], parts[1]
    start_time = parse_hhmm(start_text)
    end_time = parse_hhmm(end_text)
    if start_time is None or end_time is None:
        await update.message.reply_text("Saat formatı hatalı. Örnek: /mola 14:00,15:00")
        return
    if start_time >= end_time:
        await update.message.reply_text("Başlangıç saati bitişten küçük olmalı.")
        return

    await db_call(database.set_break_window, start_text, end_text)
    resume_time = (datetime.combine(date.today(), end_time) + timedelta(minutes=10)).strftime("%H:%M")
    await update.message.reply_text(
        f"Mola aralığı kaydedildi: {start_text}-{end_text}. Kontrol {resume_time} itibarıyla devam eder."
    )


async def sildepartman_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    name = update.message.text.partition(" ")[2].strip()
    if not name:
        await update.message.reply_text("Kullanım: /sildepartman departman_adı")
        return
    ok = await db_call(database.remove_department, name)
    await update.message.reply_text("Departman silindi." if ok else "Departman bulunamadı.")


async def yukle_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    await update.message.reply_text(
        "Kullanım:\n"
        "Excel dosyasını bu gruba /yukle açıklamasıyla gönderin.\n"
        "Sütunlar: A=personel @, B=sorumlu @, C=departman"
    )


async def yukle_document_handler(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not update.message or not update.message.document:
        return

    if not await require_auth(update):
        return

    caption = (update.message.caption or "").strip().lower()
    if not caption.startswith("/yukle"):
        return

    document = update.message.document
    file_name = document.file_name or ""
    if not file_name:
        await update.message.reply_text("Dosya adı okunamadı. Lütfen .xlsx dosyası gönderin.")
        return
    if not _is_supported_excel_file(file_name):
        await update.message.reply_text("Sadece .xlsx veya .xlsm dosyaları destekleniyor.")
        return

    telegram_file = await document.get_file()
    temp_dir = Path(tempfile.gettempdir())
    temp_path = temp_dir / f"lastseen_upload_{update.update_id}_{Path(file_name).name}"

    added_count = 0
    skipped_empty = 0
    skipped_invalid_rows: list[int] = []

    await telegram_file.download_to_drive(custom_path=str(temp_path))

    workbook = None
    try:
        workbook = load_workbook(filename=str(temp_path), read_only=True, data_only=True)
        worksheet = workbook.active

        for row_index, row in enumerate(
            worksheet.iter_rows(min_row=1, max_col=3, values_only=True),
            start=1,
        ):
            personnel = _to_cell_text(row[0] if len(row) > 0 else None)
            responsible = _to_cell_text(row[1] if len(row) > 1 else None)
            department = _to_cell_text(row[2] if len(row) > 2 else None)

            if row_index == 1 and _is_header_row(personnel, responsible, department):
                continue

            if not personnel and not responsible and not department:
                skipped_empty += 1
                continue

            if not personnel or not responsible or not department:
                skipped_invalid_rows.append(row_index)
                continue

            await db_call(database.add_personnel, personnel, responsible, department)
            added_count += 1

        invalid_preview = ", ".join(str(x) for x in skipped_invalid_rows[:10])
        invalid_suffix = "" if len(skipped_invalid_rows) <= 10 else " ..."

        summary = (
            f"Excel yüklendi.\n"
            f"Eklenen/güncellenen personel: {added_count}\n"
            f"Boş geçilen satır: {skipped_empty}\n"
            f"Hatalı satır: {len(skipped_invalid_rows)}"
        )

        if skipped_invalid_rows:
            summary += f"\nHatalı satır no: {invalid_preview}{invalid_suffix}"

        await update.message.reply_text(summary)
    except Exception as exc:
        logger.exception("Excel yükleme hatası")
        await update.message.reply_text(f"Excel işlenemedi: {type(exc).__name__}")
    finally:
        if workbook is not None:
            workbook.close()
        try:
            if temp_path.exists():
                temp_path.unlink()
        except OSError:
            pass


async def listele_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return
    records = await db_call(database.list_personnel)
    if not records:
        await update.message.reply_text("Kayıtlı personel yok.")
        return

    by_department: dict[str, dict[str, set[str]]] = {}
    dep_cache: dict[int, list[str]] = {}

    for r in records:
        department_name = r["department_name"] or "-"
        if department_name not in by_department:
            by_department[department_name] = {"responsibles": set(), "personnel": set()}

        by_department[department_name]["personnel"].add(f"@{r['username']}")

        responsible_username = r["responsible_username"]
        if responsible_username:
            by_department[department_name]["responsibles"].add(f"@{responsible_username}")

        department_id = r["department_id"]
        if department_id is not None:
            dep_id = int(department_id)
            if dep_id not in dep_cache:
                dep_cache[dep_id] = await db_call(database.get_department_responsibles, dep_id)
            for dep_responsible in dep_cache[dep_id]:
                by_department[department_name]["responsibles"].add(f"@{dep_responsible}")

    sections = []
    for department_name in sorted(by_department.keys()):
        data = by_department[department_name]
        responsibles = sorted(data["responsibles"]) if data["responsibles"] else ["-"]
        personnel = sorted(data["personnel"]) if data["personnel"] else ["-"]

        section = (
            f"Departman adı: {department_name}\n"
            f"Sorumlu: {', '.join(responsibles)}\n\n"
            f"Personeller\n"
            + "\n".join(personnel)
        )
        sections.append(section)

    await update.message.reply_text("\n\n".join(sections))


async def rapor_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not await require_auth(update):
        return

    department_query = update.message.text.partition(" ")[2].strip()
    if not department_query:
        await update.message.reply_text("Kullanım: /rapor satısekibi1")
        return

    all_personnel = await db_call(database.list_personnel)
    department_rows = [
        row
        for row in all_personnel
        if (row["department_name"] or "").casefold() == department_query.casefold()
    ]

    if not department_rows:
        await update.message.reply_text("Departman bulunamadı veya bu departmanda personel yok.")
        return

    department_name = department_rows[0]["department_name"] or department_query

    responsibles: set[str] = set()
    department_id = department_rows[0]["department_id"]
    if department_id is not None:
        for username in await db_call(database.get_department_responsibles, int(department_id)):
            responsibles.add(f"@{username}")

    for row in department_rows:
        direct_responsible = row["responsible_username"]
        if direct_responsible:
            responsibles.add(f"@{direct_responsible}")

    today_iso = get_today_local_iso()
    daily_rows = await db_call(database.get_daily_violation_counts, today_iso)
    department_daily_rows = [
        row
        for row in daily_rows
        if (row["department_name"] or "").casefold() == department_name.casefold()
    ]

    if not department_daily_rows:
        report_text = (
            f"Departman : {department_name}\n"
            f"Sorumlu : {', '.join(sorted(responsibles)) if responsibles else '-'}\n"
            "Personeller ;\n"
            "Bugün kural ihlali yok."
        )
        await update.message.reply_text(report_text)
        return

    person_counts = sorted(
        [
            (f"@{row['personnel_username']}", int(row["violation_count"]))
            for row in department_daily_rows
        ],
        key=lambda item: (-item[1], item[0]),
    )

    person_lines = [f"{username} {count} kez" for username, count in person_counts]
    report_text = (
        f"Departman : {department_name}\n"
        f"Sorumlu : {', '.join(sorted(responsibles)) if responsibles else '-'}\n"
        "Personeller ;\n"
        + "\n".join(person_lines)
    )
    await update.message.reply_text(report_text)


def _minutes_since(dt: datetime) -> int:
    now = datetime.now(timezone.utc)
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return max(0, int((now - dt).total_seconds() // 60))


async def resolve_last_seen_minutes(username: str) -> tuple[int | None, str]:
    if telethon_client is None:
        return None, "Telethon kapalı"
    try:
        user = await telethon_client.get_entity(username)
        status = getattr(user, "status", None)
        if isinstance(status, UserStatusOnline):
            return 0, "çevrimiçi"
        if isinstance(status, UserStatusOffline):
            if status.was_online:
                mins = _minutes_since(status.was_online)
                return mins, f"{mins} dakika"
            return None, "çevrimdışı (zaman bilgisi yok)"
        if isinstance(status, UserStatusRecently):
            return None, "yakınlarda"
        if isinstance(status, UserStatusLastWeek):
            return None, "son 1 hafta içinde"
        if isinstance(status, UserStatusLastMonth):
            return None, "son 1 ay içinde"
        if isinstance(status, UserStatusEmpty):
            return None, "gizli"
        return None, "bilinmiyor"
    except Exception as exc:
        logger.exception("last seen alınamadı: %s", username)
        return None, f"hata: {type(exc).__name__}"


def _should_notify_again(last_notified_at: str | None, cooldown_minutes: int) -> bool:
    if not last_notified_at:
        return True
    try:
        last = datetime.fromisoformat(last_notified_at)
    except ValueError:
        return True
    if last.tzinfo is None:
        last = last.replace(tzinfo=timezone.utc)
    return _minutes_since(last) >= cooldown_minutes


def should_notify_non_numeric_status(status_text: str) -> bool:
    informative_statuses = {
        "yakınlarda",
        "son 1 hafta içinde",
        "son 1 ay içinde",
        "çevrimdışı (zaman bilgisi yok)",
        "gizli",
        "bilinmiyor",
    }
    return status_text in informative_statuses or status_text.startswith("hata:")


async def monitor_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    if not is_within_monitor_hours():
        logger.info("Monitor atlandı: çalışma saatleri dışında.")
        return

    break_start_hhmm, break_end_hhmm = await db_call(database.get_break_window)
    if break_start_hhmm and break_end_hhmm:
        now_local = get_now_local()
        if should_skip_for_break_window(now_local, break_start_hhmm, break_end_hhmm):
            logger.info(
                "Monitor atlandı: mola aralığı aktif (%s-%s).",
                break_start_hhmm,
                break_end_hhmm,
            )
            return

    records = await db_call(database.list_personnel)
    if not records:
        logger.info("Monitor atlandı: aktif personel yok.")
        return

    target_chat_id = resolve_target_chat_id()
    if target_chat_id == 0:
        logger.warning("ALERT_CHAT_ID ayarlı değil, bildirim atlanıyor.")
        return

    today_name = today_weekday_tr()
    total_count = len(records)
    skipped_weekly_off = 0
    skipped_day_off = 0
    skipped_hourly_off = 0
    numeric_alerts = 0
    non_numeric_alerts = 0
    below_threshold = 0
    unresolved_status = 0

    for r in records:
        personnel_id = int(r["id"])
        username = r["username"]
        dep_threshold = r["department_threshold_minutes"]
        threshold = int(dep_threshold) if dep_threshold is not None else DEFAULT_THRESHOLD_MINUTES
        responsible = r["responsible_username"]
        department_name = r["department_name"]
        department = department_name or "-"

        if should_skip_for_department_weekly_off(department_name, r["department_weekly_off_day"], today_name):
            skipped_weekly_off += 1
            continue

        day_off_date = r["day_off_date"]
        if day_off_date and day_off_date == get_today_local_iso():
            skipped_day_off += 1
            continue

        exempt_until = r["exempt_until"]
        if exempt_until:
            try:
                until_dt = datetime.fromisoformat(exempt_until)
                if until_dt.tzinfo is None:
                    until_dt = until_dt.replace(tzinfo=timezone.utc)
                if until_dt > datetime.now(timezone.utc):
                    skipped_hourly_off += 1
                    continue
            except ValueError:
                pass

        mins, status_text = await resolve_last_seen_minutes(username)
        state = await db_call(database.get_watch_state, personnel_id)
        last_notified_at = state["last_notified_at"] if state else None
        status_changed = state is not None and state["last_status_text"] != status_text

        if mins is not None and mins > threshold:
            should_notify = state is None or not state["is_alerting"] or _should_notify_again(last_notified_at, ALERT_COOLDOWN_MINUTES)

            if should_notify:
                department_mentions = ""
                if r["department_id"] is not None:
                    dep_responsibles = await db_call(database.get_department_responsibles, int(r["department_id"]))
                    if dep_responsibles:
                        department_mentions = "\nDepartman sorumluları : " + " ".join(f"@{u}" for u in dep_responsibles)
                message = (
                    f"Personel : @{username}\n"
                    f"Son görülme : {mins} dakika\n"
                    f"Sorumlu : {format_responsible(responsible)}\n"
                    f"Departman : {department}"
                    f"{department_mentions}"
                )
                await context.bot.send_message(chat_id=target_chat_id, text=message)
                numeric_alerts += 1
                now_local = get_now_local()
                await db_call(
                    database.add_violation_event,
                    personnel_id=personnel_id,
                    minutes=mins,
                    occurred_at_iso=now_local.isoformat(),
                    occurred_date=now_local.date().isoformat(),
                )
                await db_call(
                    database.set_watch_state,
                    personnel_id=personnel_id,
                    is_alerting=True,
                    last_notified_at=datetime.now(timezone.utc).isoformat(),
                    last_minutes=mins,
                    last_status_text=status_text,
                )
        elif mins is None and should_notify_non_numeric_status(status_text):
            should_notify = state is None or not state["is_alerting"] or status_changed or _should_notify_again(last_notified_at, ALERT_COOLDOWN_MINUTES)

            if should_notify:
                department_mentions = ""
                if r["department_id"] is not None:
                    dep_responsibles = await db_call(database.get_department_responsibles, int(r["department_id"]))
                    if dep_responsibles:
                        department_mentions = "\nDepartman sorumluları : " + " ".join(f"@{u}" for u in dep_responsibles)
                message = (
                    f"Personel : @{username}\n"
                    f"Son görülme : {status_text} (dakika bilgisi yok)\n"
                    f"Sorumlu : {format_responsible(responsible)}\n"
                    f"Departman : {department}"
                    f"{department_mentions}"
                )
                await context.bot.send_message(chat_id=target_chat_id, text=message)
                non_numeric_alerts += 1
                await db_call(
                    database.set_watch_state,
                    personnel_id=personnel_id,
                    is_alerting=True,
                    last_notified_at=datetime.now(timezone.utc).isoformat(),
                    last_minutes=None,
                    last_status_text=status_text,
                )
        else:
            if mins is not None and mins <= threshold:
                below_threshold += 1
            elif mins is None:
                unresolved_status += 1
            await db_call(
                database.set_watch_state,
                personnel_id=personnel_id,
                is_alerting=False,
                last_notified_at=None,
                last_minutes=mins,
                last_status_text=status_text,
            )

    logger.info(
        "Monitor özeti: toplam=%d, sayisal_alarm=%d, n/a_alarm=%d, weekly_skip=%d, gunluk_izin_skip=%d, saatlik_izin_skip=%d, esik_alti=%d, cozulmeyen_durum=%d",
        total_count,
        numeric_alerts,
        non_numeric_alerts,
        skipped_weekly_off,
        skipped_day_off,
        skipped_hourly_off,
        below_threshold,
        unresolved_status,
    )


async def daily_summary_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    target_chat_id = resolve_target_chat_id()
    if target_chat_id == 0:
        logger.warning("ALERT_CHAT_ID ayarlı değil, gün sonu raporu atlanıyor.")
        return

    today_iso = get_today_local_iso()
    rows = await db_call(database.get_daily_violation_counts, today_iso)
    if not rows:
        await context.bot.send_message(
            chat_id=target_chat_id,
            text="19:30 - Gün Sonu Kural İhlalleri\n\nBugün kural ihlali yok.",
        )
        return

    grouped: dict[str, dict[str, object]] = {}
    dep_cache: dict[int, list[str]] = {}

    for row in rows:
        dep_name = row["department_name"] or "-"
        if dep_name not in grouped:
            grouped[dep_name] = {"responsibles": set(), "personnel_counts": []}

        entry = grouped[dep_name]
        responsibles = entry["responsibles"]
        personnel_counts = entry["personnel_counts"]

        personnel_counts.append((f"@{row['personnel_username']}", int(row["violation_count"])))

        direct_resp = row["responsible_username"]
        if direct_resp:
            responsibles.add(f"@{direct_resp}")

        dep_id = row["department_id"]
        if dep_id is not None:
            dep_id_int = int(dep_id)
            if dep_id_int not in dep_cache:
                dep_cache[dep_id_int] = await db_call(database.get_department_responsibles, dep_id_int)
            for dep_resp in dep_cache[dep_id_int]:
                responsibles.add(f"@{dep_resp}")

    sections: list[str] = []
    for dep_name in sorted(grouped.keys()):
        info = grouped[dep_name]
        responsibles_sorted = sorted(info["responsibles"]) if info["responsibles"] else ["-"]
        personnel_lines = [f"{username} {count} kez" for username, count in info["personnel_counts"]]

        section = (
            f"Departman adı: {dep_name}\n"
            f"Sorumlu: {', '.join(responsibles_sorted)}\n"
            f"Personeller;\n"
            + "\n".join(personnel_lines)
        )
        sections.append(section)

    message = "19:30 - Gün Sonu Kural İhlalleri\n\n" + "\n\n".join(sections)
    await context.bot.send_message(chat_id=target_chat_id, text=message)


async def daily_summary_scheduler_job(context: ContextTypes.DEFAULT_TYPE) -> None:
    now_local = get_now_local()
    report_time = time(hour=DAILY_REPORT_HOUR, minute=DAILY_REPORT_MINUTE)

    if now_local.time() < report_time:
        return

    today_iso = now_local.date().isoformat()
    last_sent_date = await db_call(database.get_app_setting, DAILY_SUMMARY_LAST_SENT_KEY)
    if last_sent_date == today_iso:
        return

    await daily_summary_job(context)
    await db_call(database.set_app_setting, DAILY_SUMMARY_LAST_SENT_KEY, today_iso)


async def init_telethon() -> None:
    global telethon_client
    if not (TELEGRAM_API_ID and TELEGRAM_API_HASH and TELETHON_STRING_SESSION):
        logger.warning("Telethon ayarları eksik. Son görülme izleme devre dışı.")
        telethon_client = None
        return

    telethon_client = TelegramClient(
        StringSession(TELETHON_STRING_SESSION),
        TELEGRAM_API_ID,
        TELEGRAM_API_HASH,
    )
    await telethon_client.connect()
    if not await telethon_client.is_user_authorized():
        logger.error("Telethon session yetkisiz. Son görülme izleme kapalı.")
        telethon_client = None
        return

    me = await telethon_client.get_me()
    if getattr(me, "bot", False):
        logger.error(
            "Telethon session bir BOT hesabına ait (%s). Son görülme için normal kullanıcı hesabı session'ı gerekli.",
            getattr(me, "username", "-"),
        )
        await telethon_client.disconnect()
        telethon_client = None


def build_app() -> Application:
    application = Application.builder().token(BOT_TOKEN).build()

    application.add_handler(CommandHandler("start", start_cmd))
    application.add_handler(CommandHandler("yardim", help_cmd))
    application.add_handler(CommandHandler("help", help_cmd))
    application.add_handler(CommandHandler("sure", sure_set_cmd))
    application.add_handler(CommandHandler("sureguncelle", sure_guncelle_cmd))
    application.add_handler(CommandHandler("personelekle", personelekle_cmd))
    application.add_handler(CommandHandler("silpersonel", silpersonel_cmd))
    application.add_handler(CommandHandler("eklesorumlu", eklesorumlu_cmd))
    application.add_handler(CommandHandler("silsorumlu", silsorumlu_cmd))
    application.add_handler(CommandHandler("ekledepartman", ekledepartman_cmd))
    application.add_handler(CommandHandler("sildepartman", sildepartman_cmd))
    application.add_handler(CommandHandler("haftalikizin", haftalikizin_cmd))
    application.add_handler(CommandHandler("kontrolhaftalikizin", kontrolhaftalikizin_cmd))
    application.add_handler(CommandHandler("izin", izin_cmd))
    application.add_handler(CommandHandler("kontrolizin", kontrolizin_cmd))
    application.add_handler(CommandHandler("saatlikizin", saatlikizin_cmd))
    application.add_handler(CommandHandler("iziniptal", iziniptal_cmd))
    application.add_handler(CommandHandler("saatlikiziniptal", saatlikiziniptal_cmd))
    application.add_handler(CommandHandler("mola", mola_cmd))
    application.add_handler(CommandHandler("yukle", yukle_cmd))
    application.add_handler(CommandHandler("rapor", rapor_cmd))
    application.add_handler(CommandHandler("listele", listele_cmd))
    application.add_handler(CommandHandler("chatid", chatid_cmd))
    application.add_handler(MessageHandler(filters.Document.ALL, yukle_document_handler))

    application.job_queue.run_repeating(
        monitor_job,
        interval=CHECK_INTERVAL_SECONDS,
        first=10,
        name="last-seen-monitor",
    )
    application.job_queue.run_repeating(
        daily_summary_scheduler_job,
        interval=DAILY_REPORT_CHECK_INTERVAL_SECONDS,
        first=15,
        name="daily-violation-summary-checker",
    )
    return application


async def main() -> None:
    validate_config()

    await db_call(database.init_db)
    await init_telethon()

    app = build_app()
    await app.initialize()
    await app.start()
    await app.updater.start_polling()
    logger.info("Bot başlatıldı.")

    try:
        while True:
            await asyncio.sleep(3600)
    finally:
        await app.updater.stop()
        await app.stop()
        await app.shutdown()
        if telethon_client is not None:
            await telethon_client.disconnect()


if __name__ == "__main__":
    asyncio.run(main())
